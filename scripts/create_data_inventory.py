#!/usr/bin/env python3
"""
生成数据清单 Excel 表格
整理所有支持的数据类型、说明和供应商信息
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 定义字段级别的数据清单（每个字段一行）
field_level_data = [
    # MatchBasicData 字段
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "match_id", "字段说明": "比赛唯一标识 ID", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "12345"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "season_id", "字段说明": "赛季唯一标识 ID", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "2023"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "competition_name", "字段说明": "联赛/杯赛名称", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "Premier League"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "home_team_id", "字段说明": "主队唯一标识 ID", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "101"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "away_team_id", "字段说明": "客队唯一标识 ID", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "102"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "home_team_name", "字段说明": "主队名称", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "Manchester City"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "away_team_name", "字段说明": "客队名称", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "Liverpool"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "match_time", "字段说明": "比赛时间（datetime 对象）", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "2024-01-15 15:00:00"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "date_unix", "字段说明": "比赛时间戳（Unix 时间）", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "1705330800"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "status", "字段说明": "比赛状态（incomplete/inprogress/complete）", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "complete"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "home_score", "字段说明": "主队最终比分", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "2"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "away_score", "字段说明": "客队最终比分", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "1"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "half_time_home", "字段说明": "主队半场比分", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "1"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "half_time_away", "字段说明": "客队半场比分", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "0"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "game_week", "字段说明": "比赛轮次", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "21"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "round_id", "字段说明": "轮次 ID", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "210"},
    {"数据大类": "MatchBasicData", "数据子类": "比赛基础数据", "字段名": "venue", "字段说明": "比赛场地", "供应商": "Footystats", "用途": "比赛列表、赛程展示", "示例值": "Etihad Stadium"},
    
    # MatchStatsData 字段
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_shots_on_target", "字段说明": "主队射正次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "6"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_shots_on_target", "字段说明": "客队射正次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "4"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_shots_off_target", "字段说明": "主队射偏次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "3"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_shots_off_target", "字段说明": "客队射偏次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "5"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_total_shots", "字段说明": "主队总射门次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "9"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_total_shots", "字段说明": "客队总射门次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "9"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_possession", "字段说明": "主队控球率（%）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "58"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_possession", "字段说明": "客队控球率（%）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "42"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_corners", "字段说明": "主队角球数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "7"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_corners", "字段说明": "客队角球数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "3"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "total_corners", "字段说明": "总角球数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "10"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_offsides", "字段说明": "主队越位次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "2"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_offsides", "字段说明": "客队越位次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "1"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_fouls", "字段说明": "主队犯规次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "11"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_fouls", "字段说明": "客队犯规次数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "9"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_yellow_cards", "字段说明": "主队黄牌数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "2"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_yellow_cards", "字段说明": "客队黄牌数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "1"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "home_red_cards", "字段说明": "主队红牌数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "0"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "away_red_cards", "字段说明": "客队红牌数", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "0"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "btts", "字段说明": "双方是否都进球（Boolean）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "True"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "over_15", "字段说明": "总进球是否超过 1.5（Boolean）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "True"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "over_25", "字段说明": "总进球是否超过 2.5（Boolean）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "True"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "over_35", "字段说明": "总进球是否超过 3.5（Boolean）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "False"},
    {"数据大类": "MatchStatsData", "数据子类": "比赛统计数据", "字段名": "winning_team_id", "字段说明": "获胜球队 ID（平局为 None）", "供应商": "Footystats", "用途": "赛后统计分析", "示例值": "101"},
    
    # MatchAdvancedData 字段
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_xg", "字段说明": "主队预期进球（xG）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.8"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_xg", "字段说明": "客队预期进球（xG）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.2"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "total_xg", "字段说明": "总预期进球", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "3.0"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_xg_prematch", "字段说明": "主队赛前预期进球", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.6"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_xg_prematch", "字段说明": "客队赛前预期进球", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.1"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "total_xg_prematch", "字段说明": "赛前总预期进球", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "2.7"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_attacks", "字段说明": "主队进攻次数", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "85"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_attacks", "字段说明": "客队进攻次数", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "62"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_dangerous_attacks", "字段说明": "主队危险进攻次数", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "42"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_dangerous_attacks", "字段说明": "客队危险进攻次数", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "31"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_lineup", "字段说明": "主队首发阵容（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "[LineupPlayer, ...]"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_lineup", "字段说明": "客队首发阵容（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "[LineupPlayer, ...]"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_bench", "字段说明": "主队替补阵容（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "[Dict, ...]"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_bench", "字段说明": "客队替补阵容（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "[Dict, ...]"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "h2h_summary", "字段说明": "交锋记录摘要（字典）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "{'total': 5, 'home_wins': 2, ...}"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_trends", "字段说明": "主队趋势分析（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "['W', 'W', 'D', 'L', 'W']"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_trends", "字段说明": "客队趋势分析（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "['L', 'D', 'W', 'L', 'D']"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "weather", "字段说明": "天气信息（字典）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "{'temp': 18, 'condition': 'Sunny'}"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "referee_id", "字段说明": "裁判 ID", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "5001"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "btts_potential", "字段说明": "双方进球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "65"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "btts_fhg_potential", "字段说明": "半场双方进球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "42"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "btts_2hg_potential", "字段说明": "下半场双方进球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "55"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "o25_potential", "字段说明": "大 2.5 球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "58"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "o35_potential", "字段说明": "大 3.5 球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "35"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "o45_potential", "字段说明": "大 4.5 球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "18"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "u25_potential", "字段说明": "小 2.5 球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "42"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "u35_potential", "字段说明": "小 3.5 球潜力（%）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "65"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "corners_potential", "字段说明": "角球潜力预测", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "9.5"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "avg_potential", "字段说明": "平均潜力值", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "52.3"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "pre_match_home_ppg", "字段说明": "主队赛前场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "2.1"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "pre_match_away_ppg", "字段说明": "客队赛前场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.5"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "pre_match_teamA_overall_ppg", "字段说明": "球队 A 赛前综合场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.9"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "pre_match_teamB_overall_ppg", "字段说明": "球队 B 赛前综合场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.6"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_ppg", "字段说明": "主队场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "2.2"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_ppg", "字段说明": "客队场均积分", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "1.4"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "h2h_previous_matches", "字段说明": "历史交锋记录（列表）", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "[Dict, ...]"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "h2h_betting_stats", "字段说明": "历史交锋赔率统计", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "{'home_win%': 40, ...}"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "home_goals_timings", "字段说明": "主队进球时间分布", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "['15', '32', '67']"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "away_goals_timings", "字段说明": "客队进球时间分布", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "['23', '78']"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "matches_completed_minimum", "字段说明": "最小已完成比赛数", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "5"},
    {"数据大类": "MatchAdvancedData", "数据子类": "比赛高级分析数据", "字段名": "game_week", "字段说明": "比赛轮次", "供应商": "Footystats", "用途": "深度分析、预测特征", "示例值": "21"},
    
    # MatchOddsData 字段
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_home", "字段说明": "主胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.85"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_draw", "字段说明": "平局赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.40"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_away", "字段说明": "客胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "4.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "implied_prob_home", "字段说明": "主胜隐含概率（%）", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "51.2"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "implied_prob_draw", "字段说明": "平局隐含概率（%）", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "27.9"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "implied_prob_away", "字段说明": "客胜隐含概率（%）", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "20.9"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "over_25_odds", "字段说明": "大 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.75"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "under_25_odds", "字段说明": "小 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.05"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "over_35_odds", "字段说明": "大 3.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.60"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "under_35_odds", "字段说明": "小 3.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.48"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "btts_yes_odds", "字段说明": "双方进球 Yes 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.65"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "btts_no_odds", "字段说明": "双方进球 No 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "handicap", "字段说明": "让球盘口", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "-0.5"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "handicap_home_odds", "字段说明": "让球主水", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.90"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "handicap_away_odds", "字段说明": "让球客水", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.90"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_doublechance_1x", "字段说明": "主胜或平局赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.22"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_doublechance_x2", "字段说明": "平局或客胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.85"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_doublechance_12", "字段说明": "主胜或客胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.30"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_result_1", "字段说明": "半场主胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.40"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_result_x", "字段说明": "半场平局赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.10"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_result_2", "字段说明": "半场客胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "4.80"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_result_1", "字段说明": "下半场主胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_result_x", "字段说明": "下半场平局赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.30"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_result_2", "字段说明": "下半场客胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "4.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_win_to_nil_1", "字段说明": "主队零封胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_win_to_nil_2", "字段说明": "客队零封胜赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "7.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_team_a_cs_yes", "字段说明": "主队零封 Yes 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.80"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_team_a_cs_no", "字段说明": "主队零封 No 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.42"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_team_b_cs_yes", "字段说明": "客队零封 Yes 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "5.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_team_b_cs_no", "字段说明": "客队零封 No 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.12"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_over_75", "字段说明": "角球大 7.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.55"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_over_85", "字段说明": "角球大 8.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.80"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_over_95", "字段说明": "角球大 9.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.10"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_over_105", "字段说明": "角球大 10.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_over_115", "字段说明": "角球大 11.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.00"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_under_75", "字段说明": "角球小 7.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.35"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_under_85", "字段说明": "角球小 8.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.95"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_under_95", "字段说明": "角球小 9.5 赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.70"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_1", "字段说明": "角球主队多赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.75"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_x", "字段说明": "角球平局赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "8.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_corners_2", "字段说明": "角球客队多赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_over05", "字段说明": "半场大 0.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.30"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_over15", "字段说明": "半场大 1.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_over25", "字段说明": "半场大 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "4.50"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_under05", "字段说明": "半场小 0.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.40"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_under15", "字段说明": "半场小 1.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.65"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_1st_half_under25", "字段说明": "半场小 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.20"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_over05", "字段说明": "下半场大 0.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.25"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_over15", "字段说明": "下半场大 1.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "2.00"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_over25", "字段说明": "下半场大 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.80"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_under05", "字段说明": "下半场小 0.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "3.80"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_under15", "字段说明": "下半场小 1.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.75"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_2nd_half_under25", "字段说明": "下半场小 2.5 球赔率", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "1.25"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "odds_comparison", "字段说明": "多家博彩公司赔率对比", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "{'FT Result': {...}}"},
    {"数据大类": "MatchOddsData", "数据子类": "比赛赔率数据", "字段名": "pinnacle_odds", "字段说明": "Pinnacle 赔率（精明资金基准）", "供应商": "Footystats", "用途": "赔率分析、价值投注识别", "示例值": "PinnacleOdds(home=1.85, draw=3.40, away=4.20)"},
]

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "字段级别数据清单"

# 设置样式
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 写入主表头
headers = ["数据大类", "数据子类", "字段名", "字段说明", "供应商", "用途", "示例值"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# 写入字段级别数据
for row_idx, data in enumerate(field_level_data, 2):
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=data.get(header, ''))
        cell.alignment = cell_alignment
        cell.border = thin_border

# 调整列宽
column_widths = [20, 18, 25, 40, 15, 25, 35]
for col_idx, width in enumerate(column_widths, 1):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# 创建数据大类汇总工作表
ws_summary = wb.create_sheet(title="数据大类汇总")

# 汇总数据
summary_headers = ["数据大类", "数据子类", "字段数量", "供应商", "主要用途", "说明"]
summary_data = [
    ["MatchBasicData", "比赛基础数据", 17, "Footystats", "比赛列表、赛程展示", "比赛基本信息，包括球队名称、比赛时间、比分、状态等"],
    ["MatchStatsData", "比赛统计数据", 22, "Footystats", "赛后统计分析", "赛后技术统计，包括射门、控球率、角球、纪律数据等"],
    ["MatchAdvancedData", "比赛高级分析数据", 42, "Footystats", "深度分析、预测特征", "深度分析数据，包括 xG、进攻次数、潜力指标、阵容、交锋记录、趋势分析等"],
    ["MatchOddsData", "比赛赔率数据", 58, "Footystats", "赔率分析、价值投注识别", "博彩赔率数据，包括胜平负、亚洲盘、大小球、角球等多种赔率"],
    ["MatchTeamsData", "比赛球队数据", "待补充", "Footystats", "球队状态分析、预测特征", "球队状态和赛季统计，包括近期战绩、主客场表现、交锋记录等"],
    ["MatchOthersData", "比赛其他补充数据", "待补充", "Footystats", "完整信息展示、特殊分析", "补充数据，包括球员数据、裁判数据、联赛统计等"],
    ["FullMatchData", "完整比赛数据", "聚合所有", "Footystats", "ML 模型预测、特征工程", "聚合所有数据类别的完整数据，用于 ML 模型预测"],
]

for col, header in enumerate(summary_headers, 1):
    cell = ws_summary.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

for row_idx, data in enumerate(summary_data, 2):
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx, value=data[col_idx-1])
        cell.alignment = cell_alignment
        cell.border = thin_border

# 调整汇总工作表列宽
summary_widths = [20, 18, 12, 15, 30, 50]
for col_idx, width in enumerate(summary_widths, 1):
    ws_summary.column_dimensions[get_column_letter(col_idx)].width = width

# 创建统计摘要工作表
ws_stats = wb.create_sheet(title="统计摘要")

# 统计内容
stats_data = [
    ["统计项", "数量/说明"],
    ["数据大类总数", "7 个"],
    ["字段总数（已整理）", f"{len(field_level_data)} 个"],
    ["", ""],
    ["供应商", "Footystats"],
    ["", ""],
    ["各大类字段数", ""],
    ["MatchBasicData", "17 个字段"],
    ["MatchStatsData", "22 个字段"],
    ["MatchAdvancedData", "42 个字段"],
    ["MatchOddsData", "58 个字段"],
    ["", ""],
    ["总数据量估算", "~40KB/场比赛"],
    ["", ""],
    ["备注", "MatchTeamsData 和 MatchOthersData 字段待继续补充"],
]

for row_idx, row_data in enumerate(stats_data, 1):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_stats.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if row_idx == 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')

ws_stats.column_dimensions['A'].width = 25
ws_stats.column_dimensions['B'].width = 45

# 保存文件
import os
script_dir = os.path.dirname(os.path.abspath(__file__))
output_path = os.path.join(script_dir, '..', 'goalcast', 'domain', 'entities', 'data_inventory.xlsx')
wb.save(output_path)

print(f"✅ Excel 表格已生成：{output_path}")
print(f"   - 字段总数：{len(field_level_data)} 个")
print(f"   - 数据大类：7 个")
print(f"   - 供应商：Footystats")
